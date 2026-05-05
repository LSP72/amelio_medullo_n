import pandas as pd
import numpy as np
from openpyxl import load_workbook
import os
from rapidfuzz import process, fuzz


class ProcessDataLokomat:
    def __init__(self):
        pass

    def read_data_info(folder_path, name_output_file):
        """This script generated an excel with:
            - the nameof the report,
            - the name of the patient
            - and the date of the first Lokomat session.

        Parameters
        ----------
        folder_path : str
            The path to the folder containing the Excel files.
        name_output_file : str
            The name of the output file (without extension).

        Returns
        -------
        pd.DataFrame
            A DataFrame containing the extracted information.
        """

        results = []

        for file in os.listdir(folder_path):
            if file.endswith(".xlsx"):
                file_path = os.path.join(folder_path, file)

                try:
                    wb = load_workbook(file_path, read_only=True)
                    sheet = wb[wb.sheetnames[0]]

                    first_sheet_name = sheet.title
                    cell_value = sheet.cell(row=3, column=3).value

                    results.append({"file_name": file, "sheet_name": first_sheet_name, "start_date": cell_value})

                except Exception as e:
                    results.append({"file_name": file, "sheet_name": "ERROR", "start_date": str(e)})

        # Convert to DataFrame
        df = pd.DataFrame(results)

        # Save to CSV
        excel_name = name_output_file + ".xlsx"
        output_path = os.path.join(folder_path, excel_name)
        df.to_excel(output_path, index=False)

        print(f"CSV saved at: {output_path}")

        return df

    @staticmethod
    def load_and_preprocess_data(file_path):

        excel_file = pd.ExcelFile(file_path)
        sheet_name = excel_file.sheet_names[0]  # Assuming the data is in the first sheet

        data = pd.read_excel(file_path, header=[0, 1])
        data.drop(data.columns[33:], axis=1, inplace=True)
        return data, sheet_name

    @staticmethod
    def clean_data_columns(data):

        data.columns = [ProcessDataLokomat._clean_col_name(col) for col in data.columns]
        print(data.columns.tolist())

        return data

    def _clean_col_name(col):
        top, sub = col

        top = str(top).strip()
        sub = str(sub).strip()

        if top.startswith("Unnamed"):
            top = ""
        if sub.startswith("Unnamed"):
            sub = ""

        top = (
            top.replace("Distance (mètres)", "Distance_m")
            .replace("Distance (pas)", "Distance_pas")
            .replace("Durée (minutes)", "Durée_min")
            .replace("Vitesse (km/h)", "Vitesse_kmh")
            .replace("Compens poids corps (%)", "BWS_%")
            .replace("Compens poids corps (kg)", "BWS_kg")
            .replace("Force de guidage gauche (%)", "Guidage_G_%")
            .replace("Force de guidage droite (%)", "Guidage_D_%")
        )

        sub = sub.replace("max", "MAX").replace("moy", "MOY").replace("min", "MIN")

        if top and sub:
            return f"{top}_{sub}"
        elif top:
            return top
        else:
            return sub

    @staticmethod
    def merge_same_day(data):

        merged_data = (
            data.groupby("Date")
            .agg(
                {
                    "Distance_m": "sum",
                    "Distance_pas": "sum",
                    "Durée_min": "sum",
                    "Vitesse_kmh_MIN": "min",
                    "Vitesse_kmh_MAX": "max",
                    "Vitesse_kmh_MOY": "mean",
                    "BWS_%_MIN": "min",
                    "BWS_%_MAX": "max",
                    "BWS_%_MOY": "mean",
                    "BWS_kg_MIN": "min",
                    "BWS_kg_MAX": "max",
                    "BWS_kg_MOY": "mean",
                    "Guidage_G_%_MIN": "min",
                    "Guidage_G_%_MAX": "max",
                    "Guidage_G_%_MOY": "mean",
                    "Guidage_D_%_MIN": "min",
                    "Guidage_D_%_MAX": "max",
                    "Guidage_D_%_MOY": "mean",
                }
            )
            .reset_index()
        )

        print(f"Rows with same dates have been merged.")
        return merged_data

    @staticmethod
    def find_patient_id(name_input, data, name_col="names", id_col="IPP", threshold=80):
        """Find the ID of the patient.

        Parameters
        ----------
        name_input : str
            Name of the patient to find.
        data : pd.DataFrame
            DataFrame containing patient information (i.e., names and IDs).
        name_col : str, optional
            Column name for the patient names, by default "names"
        id_col : str, optional
            Column name for the patient IDs, by default "IPP"
        threshold : int, optional
            Minimum similarity score for a match, by default 80

        Returns
        -------
        int or None
            The ID of the matched patient, or None if no match is found.
        """

        # Normalize everything to lowercase for comparison
        names_list = data["names"].str.lower().tolist()
        name_input_lower = name_input.lower()
        data["IPP"] = data["IPP"].fillna(0)  # Replace NaN with -1 for matching
        data["IPP"] = data["IPP"].astype(int)

        # Find the best match
        match = process.extractOne(
            name_input_lower,
            names_list,
            scorer=fuzz.WRatio,  # Handles partial matches, inversions, etc.
            score_cutoff=threshold,
        )

        if match is None:
            print(f"No match found for '{name_input}' (threshold: {threshold})")
            return None

        matched_name_lower, score, index = match
        matched_row = data["IPP"].iloc[index]

        print(f"Matched '{name_input}' → '{matched_row}' (score: {score:.1f})")
        return matched_row

    @staticmethod
    def calculate_sessions_metrics(data):
        """Function that calculates the number of weeks and the frequency of sessions per week.
           This function assumes that the data has already been merged for same-day sessions.

        Parameters
        ----------
        data : pd.DataFrame
             DataFrame containing the infos from th report.

        Returns
        -------
        nb_of_weeks, nb_sessions_per_week: tuple
            (number of weeks, average sessions per week)
        """
        data["Date"] = pd.to_datetime(data["Date"])

        weekly_counts = data.set_index("Date").resample("W").size()
        nb_of_weeks = len(weekly_counts)
        nb_sessions_per_week = weekly_counts.mean()

        return nb_of_weeks, nb_sessions_per_week
